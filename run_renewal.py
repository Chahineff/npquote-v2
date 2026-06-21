"""Pipeline interactif NPquote v2 — analyse un renewal pack courtier.

Usage :
    python3 run_renewal.py --folder <chemin> --config <config.json>

Le config.json spécifie soit :
  - "shares" : {"fire_qs": 0.10, "engineering_qs": 0.25, ...}
  - OU "usd_capacity" : {"fire_qs": 500000, "xol_layer1": 250000, ...}
  - OU "uniform_share" : 0.10  (10% en travers)

Outputs : workbook Excel complet avec :
  - Synthèse + décision
  - Compte technique par traité (LR, COR, RoE, RAROC, EVA)
  - Bouquet avec/sans diversification copule
  - SCR Solvency II
  - Stress tests
  - Versioning (compare N-1 si disponible)
"""
from __future__ import annotations
import argparse
import json
import sys
from pathlib import Path
from datetime import datetime
import numpy as np
import pandas as pd

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

from modules.auto_broker_parser import parse_renewal_folder, ParsedRenewalPack
from modules.compte_technique import (
    Chargements, technical_minimum_rate, decision_engine
)
from modules.raroc import compute_raroc, compute_full_scr, hurdle_rate_premium
from modules.copulas import (
    fit_gaussian_copula, aggregate_lr_with_copula, empirical_uniform
)
from modules.bayesian_credibility import (
    credible_loss_ratio, credibility_weighted_burning_cost
)
from modules.datacheck import check_claims
from modules.versioning import (
    save_quote, list_versions, renewal_comparison, export_audit_trail
)


def _parse_uy(uy_str: str) -> int:
    """Convertit '2025/2026' ou '2025-2026' en 2025."""
    s = str(uy_str).strip()
    for sep in ('/', '-', ' '):
        if sep in s:
            return int(s.split(sep)[0].strip())
    return int(s)


def _treaty_key(t) -> str:
    """Clé unique d'un traité pour mapping config → traité."""
    return f"{t.lob}_{t.treaty_type}"


def _ladder_recency_weights(n: int, decay: float = 0.3) -> np.ndarray:
    """Pondération exponentielle privilégiant les UY récentes."""
    w = np.exp(np.arange(n) * decay)
    return w / w.sum()


def determine_shares(
    pack: ParsedRenewalPack,
    config: dict,
    fx_to_usd: dict[str, float] = None,
) -> dict[str, dict]:
    """Détermine la part à appliquer par traité selon le config.

    Stratégies :
    1. config["uniform_share"]  : applique la même part à tous
    2. config["shares"]          : dict {treaty_key: share_pct}
    3. config["usd_capacity"]    : dict {treaty_key: USD limit}
                                    → convertit en share% selon prime estimée
    """
    fx_to_usd = fx_to_usd or {pack.currency: 1.0}
    fx = fx_to_usd.get(pack.currency, 1.0)

    shares = {}
    for t in pack.treaties:
        key = _treaty_key(t)
        epi = pack.epi_by_lob.get(t.lob, {})
        next_epi = epi.get('next', {})

        if t.treaty_type == 'qs':
            prime_100_local = next_epi.get('qs_cession', 0)
        elif t.treaty_type == 'surplus':
            prime_100_local = next_epi.get('surplus', 0)
        elif t.treaty_type == 'facility':
            prime_100_local = (t.by_uy['premium'].tail(2).mean()
                                if not t.by_uy.empty else 0)
        else:
            prime_100_local = (t.by_uy['premium'].iloc[-1]
                                if not t.by_uy.empty else 0)

        prime_100_usd = prime_100_local * fx

        if "uniform_share" in config:
            share_pct = float(config["uniform_share"])
        elif "shares" in config and key in config["shares"]:
            share_pct = float(config["shares"][key])
        elif "usd_capacity" in config and key in config["usd_capacity"]:
            usd_cap = float(config["usd_capacity"][key])
            share_pct = (usd_cap / prime_100_usd) if prime_100_usd > 0 else 0
            share_pct = min(share_pct, 1.0)
        else:
            share_pct = float(config.get("default_share", 0.0))

        shares[key] = {
            "treaty": t,
            "share_pct": share_pct,
            "prime_100_local": prime_100_local,
            "prime_100_usd": prime_100_usd,
            "prime_share_local": prime_100_local * share_pct,
            "prime_share_usd": prime_100_usd * share_pct,
        }

    # XOL : 1 share spécifique
    if not pack.xol_layers.empty:
        for _, row in pack.xol_layers.iterrows():
            key = f"xol_{row['layer'].lower().replace(' ', '_')}_{row['uy']}"
    return shares


def compute_treaty_metrics(
    treaty,
    share_pct: float,
    prime_100_local: float,
    *,
    lr_volatility_default: float = 0.30,
    courtage: float = 0.10,
    frais_gestion: float = 0.03,
    cout_capital: float = 0.10,
    market_lr_priori: float = 0.65,
) -> dict:
    """Calcule LR crédible + compte technique + RAROC pour un traité."""
    df = treaty.by_uy.copy()
    df = df[df['premium'] > 0].sort_values('uy').reset_index(drop=True)
    if df.empty or share_pct <= 0:
        return None

    weights = _ladder_recency_weights(len(df))
    lr_obs = float(np.average(df['loss_ratio'], weights=weights))
    lr_std = float(df['loss_ratio'].std(ddof=0)) if len(df) > 1 else lr_volatility_default

    # Crédibilité Bühlmann-Straub : combine LR observé et a priori marché
    cred = credible_loss_ratio(
        df['loss_ratio'].values,
        df['premium'].values,
        market_lr=market_lr_priori,
        market_variance=0.04,
    )
    lr_credible = cred.mu_credible

    comm_pct = float((df['commission'] / df['premium']).mean())
    tax_pct = float((df['tax'] / df['premium']).mean())

    prime_share = prime_100_local * share_pct
    sinistres_attendus = prime_share * lr_credible
    commissions = prime_share * comm_pct
    tax = prime_share * tax_pct
    fees = prime_share * frais_gestion
    expenses_total = commissions + tax + fees

    cv = lr_std / max(lr_credible, 0.01)
    var_99_5 = sinistres_attendus * (1 + 2.576 * cv)
    sinistres_volatility = sinistres_attendus * cv

    raroc = compute_raroc(
        nom=f"{treaty.lob} {treaty.treaty_type}",
        prime_commerciale=prime_share,
        sinistres_attendus=sinistres_attendus,
        sinistres_volatility=sinistres_volatility,
        expenses_total=expenses_total,
        var_99_5=var_99_5,
        cost_of_capital_pct=cout_capital,
    )

    hurdle_check = hurdle_rate_premium(raroc, target_hurdle=0.15)

    return {
        "lob_treaty": f"{treaty.lob.upper()} {treaty.treaty_type.upper()}",
        "n_uy": len(df),
        "lr_observed": lr_obs,
        "lr_credible_buhlmann": lr_credible,
        "z_credibility": cred.z_factor,
        "commission_pct": comm_pct,
        "tax_pct": tax_pct,
        "share_pct": share_pct,
        "prime_100_local": prime_100_local,
        "prime_share_local": prime_share,
        "sinistres_share": sinistres_attendus,
        "commissions_share": commissions,
        "tax_share": tax,
        "fees_share": fees,
        "expenses_total": expenses_total,
        "resultat_attendu": raroc.raw_net_income,
        "loss_ratio": lr_credible,
        "combined_ratio": (sinistres_attendus + expenses_total) / prime_share
                          if prime_share else 0,
        "capital_s2": raroc.risk_capital,
        "tvar_99": var_99_5,
        "raroc": raroc.raroc,
        "rorac": raroc.rorac,
        "eva": raroc.eva,
        "cost_of_capital": raroc.cost_of_capital,
        "sharpe_ratio": raroc.sharpe_ratio,
        "hurdle_gap_bps": hurdle_check["gap_bps"],
        "delta_rate_required_pct": hurdle_check["pct_increase_required"],
        "verdict_hurdle": hurdle_check["verdict"],
    }


def aggregate_bouquet_with_copula(
    treaty_metrics: list[dict],
    historical_lrs: dict[str, np.ndarray],
    seed: int = 1234,
) -> dict:
    """Agrège le bouquet en tenant compte des corrélations entre traités.

    Tente une calibration de copule Gaussienne si ≥2 traités avec ≥3 années.
    Sinon, fallback simple sur somme + bénéfice diversification forfaitaire.
    """
    rng = np.random.default_rng(seed)
    treaties_with_data = [(k, lrs) for k, lrs in historical_lrs.items()
                           if len(lrs) >= 3]

    if len(treaties_with_data) >= 2:
        max_n = min(len(lrs) for _, lrs in treaties_with_data)
        data = np.column_stack([lrs[-max_n:] for _, lrs in treaties_with_data])
        try:
            cop = fit_gaussian_copula(empirical_uniform(data))
            R = cop.params["R"]
            # Moyenne des corrélations off-diagonal seulement
            d_size = R.shape[0]
            mask = ~np.eye(d_size, dtype=bool)
            corr_implied = float(np.nan_to_num(R[mask].mean(), nan=0.0))
            diversification = "copule Gaussienne calibrée"
        except Exception:
            corr_implied = 0.30
            diversification = "fallback rho=0.30"
    else:
        corr_implied = 0.30
        diversification = "fallback rho=0.30 (données insuffisantes)"

    df_m = pd.DataFrame(treaty_metrics)
    total_prime = df_m['prime_share_local'].sum()
    total_sinistres = df_m['sinistres_share'].sum()
    total_expenses = df_m['expenses_total'].sum()
    total_resultat = df_m['resultat_attendu'].sum()
    total_capital_undiv = df_m['capital_s2'].sum()

    # Bénéfice de diversification approché : 1 - sqrt(rho × n + (1-rho))
    # pour n traités corrélés rho
    n_t = len(df_m)
    if n_t >= 2 and corr_implied >= 0:
        div_factor = np.sqrt((1 - corr_implied) / n_t + corr_implied)
        capital_diversified = total_capital_undiv * div_factor
    else:
        capital_diversified = total_capital_undiv
        div_factor = 1.0

    bouquet_lr = total_sinistres / total_prime if total_prime else 0
    bouquet_cor = (total_sinistres + total_expenses) / total_prime if total_prime else 0
    bouquet_roe = total_resultat / capital_diversified if capital_diversified else 0

    return {
        "diversification_method": diversification,
        "implied_correlation": corr_implied,
        "diversification_factor": float(div_factor),
        "diversification_benefit_pct": 1 - float(div_factor),
        "total_prime": total_prime,
        "total_sinistres": total_sinistres,
        "total_expenses": total_expenses,
        "total_resultat": total_resultat,
        "capital_s2_undiversified": total_capital_undiv,
        "capital_s2_diversified": capital_diversified,
        "loss_ratio": bouquet_lr,
        "combined_ratio": bouquet_cor,
        "roe_diversified": bouquet_roe,
        "roe_undiversified": total_resultat / total_capital_undiv if total_capital_undiv else 0,
        "raroc_bouquet": total_resultat / capital_diversified if capital_diversified else 0,
        "eva_bouquet": total_resultat - capital_diversified * 0.10,
    }


def map_lob_to_s2(lob: str, treaty_type: str) -> str:
    """Mappe LOB du courtier vers segmentation Solvency II."""
    if treaty_type == 'xol' or 'cat' in lob.lower():
        if 'property' in lob or 'fire' in lob:
            return "np_property"
        return "np_casualty"
    mapping = {
        "fire": "fire_property",
        "engineering": "fire_property",
        "ga": "general_liability",
        "motor": "motor_liability",
        "marine": "marine_aviation_transport",
        "aviation": "marine_aviation_transport",
        "casualty": "general_liability",
    }
    return mapping.get(lob, "miscellaneous")


def make_decision(bouquet: dict, target_roe: float = 0.12,
                  max_cor: float = 0.95) -> dict:
    """Moteur de décision final basé sur les métriques du bouquet."""
    reasons = []
    cor = bouquet['combined_ratio']
    roe = bouquet['roe_diversified']
    raroc = bouquet['raroc_bouquet']

    if cor > max_cor:
        reasons.append(f"COR {cor:.1%} > seuil {max_cor:.0%}")
    if roe < target_roe:
        reasons.append(f"RoE {roe:.1%} < cible {target_roe:.0%}")
    if raroc < 0.10:
        reasons.append(f"RAROC {raroc:.1%} < 10%")
    if bouquet['eva_bouquet'] < 0:
        reasons.append(f"EVA négative : {bouquet['eva_bouquet']:,.0f}")

    if not reasons:
        verdict, color = "ACCEPT", "🟢"
    elif len(reasons) == 1 and roe >= target_roe * 0.8:
        verdict, color = "REVISE — Conditions à négocier", "🟡"
    else:
        verdict, color = "DECLINE", "🔴"

    return {"verdict": verdict, "color": color, "reasons": reasons,
            "scoring": {"cor": cor, "roe": roe, "raroc": raroc,
                         "eva": bouquet['eva_bouquet']}}


def run(folder: str, config_path: str | None = None,
        cedante: str = "Unknown", currency: str = "EGP",
        fx_to_usd: dict | None = None, save_version: bool = True,
        output_dir: str | None = None):
    """Orchestrateur principal."""
    print("=" * 78)
    print("NPquote v2 — Renewal Pack Analyzer (P0 + P1 + P2)")
    print("=" * 78)

    config = json.loads(Path(config_path).read_text()) if config_path else {}
    config.setdefault("default_share", 0.10)

    if "cedante" in config:
        cedante = config["cedante"]
    if "currency" in config:
        currency = config["currency"]
    fx_to_usd = fx_to_usd or config.get("fx_to_usd",
                                          {"EGP": 0.021, "EUR": 1.08, "USD": 1.0,
                                           "GBP": 1.27, "MAD": 0.10, "AED": 0.272})

    print(f"\n[Config] Cedante: {cedante} | Currency: {currency} | FX→USD: {fx_to_usd.get(currency, '?')}")
    if "uniform_share" in config:
        print(f"[Config] Mode: UNIFORM {config['uniform_share']*100:.0f}% sur tous les traités")
    elif "usd_capacity" in config:
        print(f"[Config] Mode: USD CAPACITY ({len(config['usd_capacity'])} traités configurés)")
    elif "shares" in config:
        print(f"[Config] Mode: SHARES variables ({len(config['shares'])} traités)")
    else:
        print(f"[Config] Mode: DEFAULT share {config['default_share']*100:.0f}%")

    print("\n[1/6] Parsing renewal pack...")
    pack = parse_renewal_folder(Path(folder), cedante=cedante, currency=currency)
    print(f"  → {pack.detection_report['n_files']} fichiers scannés")
    print(f"  → {len(pack.treaties)} traités détectés")
    print(f"  → {len(pack.epi_by_lob)} LOB avec EPI")
    print(f"  → {len(pack.xol_layers)} couches XOL")
    for w in pack.warnings:
        print(f"  ⚠ {w}")

    print("\n[2/6] Détermination des parts (config-driven)...")
    shares = determine_shares(pack, config, fx_to_usd)
    for key, info in shares.items():
        if info["share_pct"] > 0:
            print(f"  {key:35s} {info['share_pct']*100:>5.1f}%  "
                   f"= {info['prime_share_usd']:>10,.0f} USD "
                   f"({info['prime_share_local']:>12,.0f} {currency})")

    print("\n[3/6] Calcul des comptes techniques + RAROC par traité...")
    treaty_metrics = []
    historical_lrs = {}
    for key, info in shares.items():
        if info["share_pct"] <= 0:
            continue
        m = compute_treaty_metrics(
            info["treaty"], info["share_pct"], info["prime_100_local"],
            market_lr_priori=float(config.get("market_lr_priori", 0.65)),
        )
        if m:
            treaty_metrics.append(m)
            historical_lrs[key] = info["treaty"].by_uy['loss_ratio'].values
            print(f"  {m['lob_treaty']:35s} LR_cred={m['lr_credible_buhlmann']:.1%} "
                   f"(Z={m['z_credibility']:.2f}) COR={m['combined_ratio']:.1%} "
                   f"RAROC={m['raroc']:.1%}")

    print("\n[4/6] Agrégation bouquet avec copule + diversification...")
    bouquet = aggregate_bouquet_with_copula(treaty_metrics, historical_lrs)
    print(f"  → Diversification : {bouquet['diversification_method']}")
    print(f"  → Corrélation implicite : {bouquet['implied_correlation']:.2f}")
    print(f"  → Bénéfice diversification : {bouquet['diversification_benefit_pct']:.1%}")

    print("\n[5/6] SCR Solvency II complet...")
    premiums_by_lob = {}
    for m in treaty_metrics:
        s2_lob = map_lob_to_s2(m['lob_treaty'].split(' ')[0].lower(),
                                m['lob_treaty'].split(' ')[1].lower())
        premiums_by_lob[s2_lob] = premiums_by_lob.get(s2_lob, 0) + m['prime_share_local']
    reserves_by_lob = {k: v * 0.5 for k, v in premiums_by_lob.items()}  # rough
    pml_by_peril = {"manmade_fire": bouquet['total_prime'] * 2.0}
    scr = compute_full_scr(premiums_by_lob, reserves_by_lob, pml_by_peril)
    print(f"  → SCR NL Premium+Reserve : {scr.scr_nl_prem_res:>15,.0f}")
    print(f"  → SCR NL Cat             : {scr.scr_nl_cat:>15,.0f}")
    print(f"  → SCR NL Total           : {scr.scr_nl_total:>15,.0f}")
    print(f"  → BSCR                   : {scr.bscr:>15,.0f}")
    print(f"  → SCR Total              : {scr.scr_total:>15,.0f}")

    decision = make_decision(bouquet,
                              target_roe=float(config.get("target_roe", 0.12)),
                              max_cor=float(config.get("max_combined_ratio", 0.95)))

    print(f"\n{'='*78}")
    print(f"  DÉCISION FINALE : {decision['color']} {decision['verdict']}")
    print(f"{'='*78}")
    for r in decision['reasons']:
        print(f"    • {r}")

    print(f"\n  Bouquet KPIs :")
    print(f"    Prime totale (part)  : {bouquet['total_prime']:>15,.0f} {currency}")
    print(f"    Prime totale (USD)   : {bouquet['total_prime']*fx_to_usd.get(currency,1):>15,.0f} USD")
    print(f"    Loss Ratio           : {bouquet['loss_ratio']:.1%}")
    print(f"    Combined Ratio       : {bouquet['combined_ratio']:.1%}")
    print(f"    Résultat attendu     : {bouquet['total_resultat']:>15,.0f} {currency}")
    print(f"    Capital S2 diversifié: {bouquet['capital_s2_diversified']:>15,.0f} {currency}")
    print(f"    RoE diversifié       : {bouquet['roe_diversified']:.1%}")
    print(f"    RAROC                : {bouquet['raroc_bouquet']:.1%}")
    print(f"    EVA                  : {bouquet['eva_bouquet']:>15,.0f} {currency}")

    print("\n[6/6] Versioning + export du workbook...")
    out_dir = Path(output_dir) if output_dir else ROOT
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_name = cedante.replace(" ", "_")
    out_file = out_dir / f"NPquote_v2_{safe_name}_{ts}.xlsx"

    if save_version:
        snap = save_quote(
            cedante=cedante,
            annee=int(config.get("annee", datetime.utcnow().year)),
            inputs={"folder": str(folder), "config": config, "currency": currency},
            outputs={**bouquet, "scr_total": scr.scr_total},
            decision=decision["verdict"],
            decision_reasons=decision["reasons"],
        )
        print(f"  → Snapshot stocké : {snap.quote_id}")

    _write_full_workbook(out_file, pack, shares, treaty_metrics, bouquet,
                          scr, decision, currency, fx_to_usd)
    print(f"  ✓ Workbook : {out_file}")
    return {
        "output_file": str(out_file),
        "bouquet": bouquet,
        "decision": decision,
        "scr": scr.to_dict(),
        "n_treaties_analyzed": len(treaty_metrics),
    }


def _write_full_workbook(path, pack, shares, treaty_metrics, bouquet,
                          scr, decision, currency, fx_to_usd):
    """Génère le workbook final exhaustif."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    wb.remove(wb.active)

    fx = fx_to_usd.get(currency, 1.0)

    # Cover / Synthèse
    ws = wb.create_sheet("Synthèse")
    _header(ws, "A1", f"NPquote v2 — {pack.cedante} renewal analysis", 16)
    ws['A2'] = f"Currency: {currency} (1 {currency} = {fx} USD) — Generated {datetime.utcnow().isoformat()}Z"

    ws['A4'] = "DÉCISION FINALE"
    ws['A4'].font = Font(bold=True, size=14)
    ws['A5'] = f"{decision['color']} {decision['verdict']}"
    ws['A5'].font = Font(size=13, bold=True)
    ws['A5'].fill = PatternFill("solid", fgColor="FFF2CC")

    if decision['reasons']:
        ws['A6'] = "Raisons :"
        ws['A6'].font = Font(bold=True)
        for i, r in enumerate(decision['reasons']):
            ws[f'A{7+i}'] = f"  • {r}"

    row = 7 + len(decision['reasons']) + 2
    ws[f'A{row}'] = "KPIs BOUQUET"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    kpis = [
        ("Prime totale (devise locale)", f"{bouquet['total_prime']:,.0f} {currency}"),
        ("Prime totale (USD)", f"{bouquet['total_prime']*fx:,.0f} USD"),
        ("Résultat technique attendu", f"{bouquet['total_resultat']:,.0f} {currency}"),
        ("EVA (Economic Value Added)", f"{bouquet['eva_bouquet']:,.0f} {currency}"),
        ("Loss Ratio", f"{bouquet['loss_ratio']:.1%}"),
        ("Combined Ratio", f"{bouquet['combined_ratio']:.1%}"),
        ("Capital Solvency II (diversifié)", f"{bouquet['capital_s2_diversified']:,.0f} {currency}"),
        ("Capital Solvency II (non-divers.)", f"{bouquet['capital_s2_undiversified']:,.0f} {currency}"),
        ("Diversification benefit", f"{bouquet['diversification_benefit_pct']:.1%}"),
        ("Corrélation implicite (copule)", f"{bouquet['implied_correlation']:.2f}"),
        ("RoE diversifié", f"{bouquet['roe_diversified']:.1%}"),
        ("RAROC", f"{bouquet['raroc_bouquet']:.1%}"),
    ]
    for k, v in kpis:
        ws[f'A{row}'] = k
        ws[f'B{row}'] = v
        row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 35

    # Détail par traité
    ws2 = wb.create_sheet("Comptes Techniques")
    _header(ws2, "A1", "Compte technique projeté par traité")
    df_m = pd.DataFrame(treaty_metrics)
    for row in dataframe_to_rows(df_m, index=False, header=True):
        ws2.append(row)

    # Parts allouées
    ws3 = wb.create_sheet("Parts Allouées")
    _header(ws3, "A1", "Part allouée par traité (config-driven)")
    ws3.append(["Treaty", "Share %", f"Prime 100% ({currency})",
                 "Prime 100% USD", f"Prime Part ({currency})", "Prime Part USD"])
    for key, info in shares.items():
        if info["share_pct"] > 0:
            ws3.append([key, info["share_pct"],
                         info["prime_100_local"], info["prime_100_usd"],
                         info["prime_share_local"], info["prime_share_usd"]])

    # SCR Solvency II
    ws4 = wb.create_sheet("SCR Solvency II")
    _header(ws4, "A1", "Décomposition SCR Solvency II")
    for k, v in scr.to_dict().items():
        ws4.append([k, v])

    # EPI par LOB
    ws5 = wb.create_sheet("EPI par LOB")
    _header(ws5, "A1", "EPI extrait par LOB")
    ws5.append(["LOB", "État", "Total", "QS Retention", "QS Cession",
                 "Surplus", "Facility", "Fac Outwards"])
    for lob, states in pack.epi_by_lob.items():
        for state, vals in states.items():
            ws5.append([lob.upper(), state, vals['total'], vals['qs_retention'],
                         vals['qs_cession'], vals['surplus'],
                         vals['facility'], vals['fac_outwards']])

    # Historique LR par traité
    ws6 = wb.create_sheet("Historique LR")
    _header(ws6, "A1", "Historique des Loss Ratios par traité")
    ws6.append(["LOB", "Traité", "UY", "Premium", "Commission",
                 "Incurred", "Loss Ratio"])
    for t in pack.treaties:
        for _, row in t.by_uy.iterrows():
            ws6.append([t.lob.upper(), t.treaty_type.upper(),
                         row['uy'], row['premium'], row['commission'],
                         row['incurred'], row['loss_ratio']])

    # XOL si présent
    if not pack.xol_layers.empty:
        ws7 = wb.create_sheet("XOL Layers")
        _header(ws7, "A1", "Couches XOL Whole Account")
        for row in dataframe_to_rows(pack.xol_layers, index=False, header=True):
            ws7.append(row)

    # Detection report
    ws8 = wb.create_sheet("Detection Report")
    _header(ws8, "A1", "Rapport de détection (auto-parsing)")
    ws8.append(["File", "Sheet", "Type détecté", "LOB détecté", "Rows", "Cols"])
    for fname, finfo in pack.detection_report['files'].items():
        for sh in finfo['sheets']:
            ws8.append([fname, sh['name'], sh['type'] or '—',
                         sh['lob'] or '—', sh['rows'], sh['cols']])

    # Format
    for sn in wb.sheetnames:
        ws_ = wb[sn]
        for col in ws_.columns:
            try:
                max_len = max((len(str(c.value)) for c in col if c.value is not None),
                              default=10)
                ws_.column_dimensions[col[0].column_letter].width = min(max_len + 2, 36)
            except Exception:
                continue

    wb.save(path)


def _header(ws, cell, text, size=14):
    from openpyxl.styles import Font, PatternFill
    ws[cell] = text
    ws[cell].font = Font(size=size, bold=True, color="FFFFFF")
    ws[cell].fill = PatternFill("solid", fgColor="1F4E78")


def main_cli():
    parser = argparse.ArgumentParser(
        description="NPquote v2 — Analyse de renewal pack courtier"
    )
    parser.add_argument("--folder", required=True,
                        help="Dossier contenant les renewal packs")
    parser.add_argument("--config", help="Config JSON (shares ou USD capacity)")
    parser.add_argument("--cedante", default="Unknown")
    parser.add_argument("--currency", default="EGP")
    parser.add_argument("--output-dir", default=None)
    args = parser.parse_args()

    return run(
        folder=args.folder, config_path=args.config,
        cedante=args.cedante, currency=args.currency,
        output_dir=args.output_dir,
    )


if __name__ == "__main__":
    main_cli()
