"""P2.4 — Versioning et audit trail des cotations.

Persistance JSON des cotations avec :
- Snapshot inputs / outputs / décision
- Diff entre 2 versions (changements de tarif, conditions, données)
- Historique par cédante (comparaison N-1 vs N)
- Métriques d'évolution (rate change, exposure change, LR evolution)
"""
from __future__ import annotations
import json
import hashlib
from dataclasses import dataclass, asdict, field
from pathlib import Path
from datetime import datetime
import pandas as pd


VERSIONS_DIR = Path.home() / ".npquote_v2_versions"


@dataclass
class CotationSnapshot:
    """Snapshot complet d'une cotation à un instant donné."""
    quote_id: str             # ex "AlWataniya_2026_v3"
    cedante: str
    annee: int
    timestamp: str
    underwriter: str
    inputs_hash: str           # hash des inputs pour détecter changements
    inputs: dict               # tous les inputs (EPI, sinistres résumés, structure)
    outputs: dict              # tous les outputs (taux, primes, comptes techniques)
    decision: str              # ACCEPT / REVISE / DECLINE
    decision_reasons: list = field(default_factory=list)
    notes: str = ""

    def to_dict(self):
        return asdict(self)

    @classmethod
    def from_dict(cls, d):
        return cls(**d)


def _compute_inputs_hash(inputs: dict) -> str:
    """Hash SHA-256 des inputs (canonical JSON) pour détecter modifications."""
    canonical = json.dumps(inputs, sort_keys=True, default=str)
    return hashlib.sha256(canonical.encode()).hexdigest()[:16]


def save_quote(
    cedante: str,
    annee: int,
    inputs: dict,
    outputs: dict,
    decision: str,
    *,
    underwriter: str = "system",
    decision_reasons: list = None,
    notes: str = "",
    versions_dir: Path = VERSIONS_DIR,
) -> CotationSnapshot:
    """Persiste une nouvelle version de cotation.

    Auto-versioning : v1, v2, v3, ... par (cédante, année).
    """
    versions_dir.mkdir(parents=True, exist_ok=True)
    cedante_safe = cedante.replace(" ", "_").replace("/", "_")
    pattern = f"{cedante_safe}_{annee}_v*.json"
    existing = sorted(versions_dir.glob(pattern))
    next_v = len(existing) + 1
    quote_id = f"{cedante_safe}_{annee}_v{next_v}"

    snap = CotationSnapshot(
        quote_id=quote_id, cedante=cedante, annee=annee,
        timestamp=datetime.utcnow().isoformat() + "Z",
        underwriter=underwriter,
        inputs_hash=_compute_inputs_hash(inputs),
        inputs=inputs, outputs=outputs,
        decision=decision, decision_reasons=decision_reasons or [],
        notes=notes,
    )
    path = versions_dir / f"{quote_id}.json"
    path.write_text(json.dumps(snap.to_dict(), default=str, indent=2))
    return snap


def list_versions(
    cedante: str | None = None,
    versions_dir: Path = VERSIONS_DIR,
) -> pd.DataFrame:
    """Liste toutes les versions stockées, filtrable par cédante."""
    versions_dir.mkdir(parents=True, exist_ok=True)
    pattern = f"{cedante.replace(' ', '_')}_*.json" if cedante else "*.json"
    files = sorted(versions_dir.glob(pattern))
    rows = []
    for f in files:
        try:
            d = json.loads(f.read_text())
            rows.append({
                "quote_id": d["quote_id"],
                "cedante": d["cedante"],
                "annee": d["annee"],
                "timestamp": d["timestamp"],
                "underwriter": d["underwriter"],
                "decision": d["decision"],
                "inputs_hash": d["inputs_hash"],
                "path": str(f),
            })
        except (json.JSONDecodeError, KeyError):
            continue
    return pd.DataFrame(rows)


def load_quote(quote_id: str, versions_dir: Path = VERSIONS_DIR
               ) -> CotationSnapshot | None:
    """Charge une version par son ID."""
    path = versions_dir / f"{quote_id}.json"
    if not path.exists():
        return None
    return CotationSnapshot.from_dict(json.loads(path.read_text()))


def diff_versions(quote_id_a: str, quote_id_b: str,
                  versions_dir: Path = VERSIONS_DIR) -> dict:
    """Compare 2 versions et retourne les changements numériques.

    Pour chaque champ commun de outputs :
      - valeur A, valeur B, delta, delta_pct
    """
    a = load_quote(quote_id_a, versions_dir)
    b = load_quote(quote_id_b, versions_dir)
    if a is None or b is None:
        return {"error": f"Version non trouvée : {quote_id_a if a is None else quote_id_b}"}

    changes = []
    for key in set(a.outputs.keys()) | set(b.outputs.keys()):
        va, vb = a.outputs.get(key), b.outputs.get(key)
        if isinstance(va, (int, float)) and isinstance(vb, (int, float)):
            delta = vb - va
            pct = delta / va if va != 0 else float('inf')
            changes.append({
                "field": key, "version_a": va, "version_b": vb,
                "delta": delta, "delta_pct": pct,
            })
        elif va != vb:
            changes.append({
                "field": key, "version_a": va, "version_b": vb,
                "delta": None, "delta_pct": None,
            })
    changes_df = pd.DataFrame(changes).sort_values(
        "delta_pct", key=lambda s: s.abs() if pd.api.types.is_numeric_dtype(s) else s,
        ascending=False, na_position='last',
    )

    return {
        "version_a": {"id": a.quote_id, "ts": a.timestamp, "decision": a.decision},
        "version_b": {"id": b.quote_id, "ts": b.timestamp, "decision": b.decision},
        "hash_changed": a.inputs_hash != b.inputs_hash,
        "decision_changed": a.decision != b.decision,
        "changes": changes_df,
    }


def renewal_comparison(
    cedante: str,
    current_year: int,
    versions_dir: Path = VERSIONS_DIR,
) -> dict:
    """Compare la cotation N-1 vs N pour un même client.

    Métriques clefs : rate change, exposure change, LR evolution, capital evolution.
    """
    df = list_versions(cedante, versions_dir)
    if df.empty:
        return {"error": f"Aucune cotation pour {cedante}"}

    # Sélectionne la dernière version de chaque année
    df = df.sort_values("timestamp")
    by_year = df.groupby("annee").last().reset_index()

    rows = []
    for _, row in by_year.iterrows():
        snap = load_quote(row['quote_id'], versions_dir)
        rows.append({
            "annee": int(row['annee']),
            "quote_id": row['quote_id'],
            "decision": snap.decision,
            "prime": snap.outputs.get("prime_totale", 0),
            "lr_attendu": snap.outputs.get("loss_ratio", 0),
            "combined_ratio": snap.outputs.get("combined_ratio", 0),
            "roe": snap.outputs.get("roe_technique", 0),
            "resultat": snap.outputs.get("resultat_technique", 0),
        })

    hist = pd.DataFrame(rows).sort_values("annee").reset_index(drop=True)

    if len(hist) >= 2:
        hist['rate_change_pct'] = hist['prime'].pct_change()
        hist['lr_change_bps'] = (hist['lr_attendu'].diff() * 10000)

    return {"cedante": cedante, "history": hist}


def export_audit_trail(
    cedante: str,
    output_path: Path,
    versions_dir: Path = VERSIONS_DIR,
):
    """Exporte tout l'historique d'un client en XLSX (compliance & audit)."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Audit Trail")
    ws.append(["AUDIT TRAIL — Réassurance"])
    ws.append([f"Cédante : {cedante}"])
    ws.append([f"Exporté : {datetime.utcnow().isoformat()}Z"])
    ws.append([])

    versions = list_versions(cedante, versions_dir)
    ws.append(["VERSIONS"])
    for row in dataframe_to_rows(versions, index=False, header=True):
        ws.append(row)

    if len(versions) >= 2:
        ws_diff = wb.create_sheet("Diff Latest")
        diff = diff_versions(versions.iloc[-2]['quote_id'],
                              versions.iloc[-1]['quote_id'],
                              versions_dir)
        ws_diff.append([f"Diff {diff['version_a']['id']} → {diff['version_b']['id']}"])
        ws_diff.append([])
        if 'changes' in diff and not diff['changes'].empty:
            for row in dataframe_to_rows(diff['changes'], index=False, header=True):
                ws_diff.append(row)

    wb.save(output_path)
