"""P3.1 — UW Sheet : format slip souscripteur traité professionnel.

Réplique le format standard utilisé par les souscripteurs réassurance
(slip ASR / SMGA / SCOR / Munich Re style) pour cotations Prop + Non-Prop.

Sections :
- Header  : Company, Broker, Inception, Period, Currency, FX
- PROP    : Treaty | Share | Leader | Limit @100% | Ceded Limit | Event Limit |
            EPI ORG CUR | EPI OUR SHARE USD | LIMIT OUR SHARE USD | Hist LR |
            Est LR | Commission | Brokerage
- NON-PROP: Treaty | Share | Leader | Limit @100% | Deductible | Reinstatements |
            Adj rate | GNPI | EPI ORG CUR | LIMIT OUR SHARE USD |
            EPI OUR SHARE USD | Brokerage | Reinsurance Tax
- Totals + signatures section
"""
from __future__ import annotations
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from typing import Optional
import pandas as pd
import numpy as np


@dataclass
class TreatyHeader:
    """En-tête commune du slip cotation."""
    company: str = ""              # Cédante
    broker: str = ""               # Courtier
    leader: str = "ASR"            # Leader réassureur (souvent SMGA AME, ASR, etc.)
    inception_date: str = "2026-01-01"
    expiry_date: str = "2026-12-31"
    original_currency: str = "USD"
    exchange_rate_to_usd: float = 1.0
    period_basis: str = "RAD (Risks Attaching During)"  # RAD ou LOD
    business_class: str = "All classes (treaty whole account)"
    country: str = ""
    underwriter: str = ""
    quote_id: str = ""

    def to_dict(self):
        return asdict(self)


@dataclass
class PropTreatyLine:
    """Une ligne traité proportionnel dans le slip."""
    treaty_name: str               # ex "Fire Surplus"
    share_pct: float = 0.0         # part SMGA (0..1)
    leader: str = "ASR"
    limit_org_cur_100pct: float = 0.0    # capacity totale 100%
    ceded_limit_org_cur: float = 0.0     # capacity cédée (= retention × lines)
    event_limit_org_cur: float = 0.0     # limit par événement (EML cat / single risk)
    epi_org_cur: float = 0.0             # EPI 100%
    historical_lr: float = 0.0
    estimated_lr: float = 0.0
    commission_pct: float = 0.30         # COM QS
    profit_commission_pct: float = 0.0   # PC if applicable
    brokerage_pct: float = 0.025
    sliding_scale_range: str = ""        # ex "min 25% / max 40% pivot 60%"
    notes: str = ""

    def epi_our_share_usd(self, fx: float) -> float:
        return self.epi_org_cur * self.share_pct * fx

    def limit_our_share_usd(self, fx: float) -> float:
        return self.ceded_limit_org_cur * self.share_pct * fx


@dataclass
class NonPropTreatyLine:
    """Une ligne traité non-proportionnel (XL) dans le slip."""
    treaty_name: str               # ex "Layer 1 — Whole Account"
    layer_n: int = 1
    share_pct: float = 0.0
    leader: str = "ASR"
    limit_org_cur_100pct: float = 0.0    # garantie de la couche (limit XS deductible)
    deductible_org_cur: float = 0.0      # priority / retention
    reinstatement_str: str = "3@100%"    # nb rec et taux : "1@100%" / "2@100%" / "Free"
    adjustable_rate_pct: float = 0.0     # rate on premium (% GNPI)
    gnpi_org_cur: float = 0.0            # GNPI 100%
    epi_org_cur: float = 0.0             # = adjustable_rate × GNPI (provisional)
    brokerage_pct: float = 0.10
    reinsurance_tax_pct: float = 0.0     # tax / stamp duty si applicable
    notes: str = ""

    def limit_our_share_usd(self, fx: float) -> float:
        return self.limit_org_cur_100pct * self.share_pct * fx

    def epi_our_share_usd(self, fx: float) -> float:
        return self.epi_org_cur * self.share_pct * fx


@dataclass
class UWSheet:
    """Slip souscripteur complet."""
    header: TreatyHeader
    prop_lines: list[PropTreatyLine] = field(default_factory=list)
    nonprop_lines: list[NonPropTreatyLine] = field(default_factory=list)
    clauses: list[str] = field(default_factory=list)
    exclusions: list[str] = field(default_factory=list)
    warranties: list[str] = field(default_factory=list)

    def total_prop(self) -> dict:
        """Totaux section proportionnelle."""
        fx = self.header.exchange_rate_to_usd
        lines = self.prop_lines
        total_limit_100 = sum(l.limit_org_cur_100pct for l in lines)
        total_ceded = sum(l.ceded_limit_org_cur for l in lines)
        total_epi_100 = sum(l.epi_org_cur for l in lines)
        total_epi_usd = sum(l.epi_our_share_usd(fx) for l in lines)
        total_limit_usd = sum(l.limit_our_share_usd(fx) for l in lines)
        # LR moyenne pondérée par EPI
        if total_epi_100 > 0:
            avg_hist = sum(l.historical_lr * l.epi_org_cur for l in lines) / total_epi_100
            avg_est = sum(l.estimated_lr * l.epi_org_cur for l in lines) / total_epi_100
        else:
            avg_hist = avg_est = 0
        # Brokerage moyen pondéré
        avg_brok = (sum(l.brokerage_pct * l.epi_org_cur for l in lines) / total_epi_100
                    if total_epi_100 else 0)
        return {
            "limit_100pct": total_limit_100,
            "ceded_limit": total_ceded,
            "epi_100pct": total_epi_100,
            "epi_our_share_usd": total_epi_usd,
            "limit_our_share_usd": total_limit_usd,
            "historical_lr_avg": avg_hist,
            "estimated_lr_avg": avg_est,
            "brokerage_avg": avg_brok,
        }

    def total_nonprop(self) -> dict:
        """Totaux section non-proportionnelle."""
        fx = self.header.exchange_rate_to_usd
        lines = self.nonprop_lines
        total_limit_100 = sum(l.limit_org_cur_100pct for l in lines)
        min_ded = min((l.deductible_org_cur for l in lines), default=0)
        sum_rate = sum(l.adjustable_rate_pct for l in lines)
        total_gnpi = max((l.gnpi_org_cur for l in lines), default=0)
        total_epi_100 = sum(l.epi_org_cur for l in lines)
        total_limit_usd = sum(l.limit_our_share_usd(fx) for l in lines)
        total_epi_usd = sum(l.epi_our_share_usd(fx) for l in lines)
        avg_brok = (sum(l.brokerage_pct * l.epi_org_cur for l in lines) / total_epi_100
                    if total_epi_100 else 0)
        return {
            "limit_100pct": total_limit_100,
            "min_deductible": min_ded,
            "sum_rate": sum_rate,
            "max_gnpi": total_gnpi,
            "epi_100pct": total_epi_100,
            "limit_our_share_usd": total_limit_usd,
            "epi_our_share_usd": total_epi_usd,
            "brokerage_avg": avg_brok,
        }

    def grand_total(self) -> dict:
        """Récapitulatif total des deux sections."""
        p = self.total_prop()
        np_ = self.total_nonprop()
        return {
            "total_epi_our_share_usd": p["epi_our_share_usd"] + np_["epi_our_share_usd"],
            "total_limit_our_share_usd": p["limit_our_share_usd"] + np_["limit_our_share_usd"],
            "expected_loss_usd_prop": p["epi_our_share_usd"] * p["estimated_lr_avg"],
            "expected_loss_usd_nonprop": np_["epi_our_share_usd"] * 0.40,  # default LR XL
            "expected_total_loss_usd":
                p["epi_our_share_usd"] * p["estimated_lr_avg"]
                + np_["epi_our_share_usd"] * 0.40,
        }


def prop_to_dataframe(sheet: UWSheet) -> pd.DataFrame:
    """Convertit la section PROP en DataFrame pour Excel/affichage."""
    fx = sheet.header.exchange_rate_to_usd
    cur = sheet.header.original_currency
    rows = []
    for l in sheet.prop_lines:
        rows.append({
            "Treaty": l.treaty_name,
            "Share": l.share_pct,
            "Leader": l.leader,
            f"Limit in {cur} @100%": l.limit_org_cur_100pct,
            f"Ceded Limit in {cur}": l.ceded_limit_org_cur,
            f"Event Limit in {cur}": l.event_limit_org_cur or "",
            f"EPI in {cur}": l.epi_org_cur,
            "EPI OUR SHARE USD": l.epi_our_share_usd(fx),
            "LIMIT OUR SHARE USD": l.limit_our_share_usd(fx),
            "Historical LR": l.historical_lr,
            "Estimated LR": l.estimated_lr,
            "Commission QS": l.commission_pct,
            "Profit Comm": l.profit_commission_pct,
            "Brokerage": l.brokerage_pct,
            "Sliding Scale": l.sliding_scale_range,
            "Notes": l.notes,
        })
    if rows:
        t = sheet.total_prop()
        rows.append({
            "Treaty": "TOTAL PROP", "Share": "", "Leader": "",
            f"Limit in {cur} @100%": t["limit_100pct"],
            f"Ceded Limit in {cur}": t["ceded_limit"],
            f"Event Limit in {cur}": "",
            f"EPI in {cur}": t["epi_100pct"],
            "EPI OUR SHARE USD": t["epi_our_share_usd"],
            "LIMIT OUR SHARE USD": t["limit_our_share_usd"],
            "Historical LR": t["historical_lr_avg"],
            "Estimated LR": t["estimated_lr_avg"],
            "Commission QS": "", "Profit Comm": "",
            "Brokerage": t["brokerage_avg"], "Sliding Scale": "", "Notes": "",
        })
    return pd.DataFrame(rows)


def nonprop_to_dataframe(sheet: UWSheet) -> pd.DataFrame:
    """Convertit la section NON-PROP en DataFrame."""
    fx = sheet.header.exchange_rate_to_usd
    cur = sheet.header.original_currency
    rows = []
    for l in sheet.nonprop_lines:
        rows.append({
            "Treaty / Layer": l.treaty_name,
            "Share": l.share_pct,
            "Leader": l.leader,
            f"Limit in {cur} @100%": l.limit_org_cur_100pct,
            f"Deductible in {cur}": l.deductible_org_cur,
            "Reinstatement": l.reinstatement_str,
            "Adjustable Rate %": l.adjustable_rate_pct,
            f"GNPI in {cur}": l.gnpi_org_cur,
            f"EPI in {cur}": l.epi_org_cur,
            "LIMIT OUR SHARE USD": l.limit_our_share_usd(fx),
            "EPI OUR SHARE USD": l.epi_our_share_usd(fx),
            "Brokerage": l.brokerage_pct,
            "Reins Tax": l.reinsurance_tax_pct,
            "Notes": l.notes,
        })
    if rows:
        t = sheet.total_nonprop()
        rows.append({
            "Treaty / Layer": "TOTAL NON PROP", "Share": "", "Leader": "",
            f"Limit in {cur} @100%": t["limit_100pct"],
            f"Deductible in {cur}": t["min_deductible"],
            "Reinstatement": "",
            "Adjustable Rate %": t["sum_rate"],
            f"GNPI in {cur}": t["max_gnpi"],
            f"EPI in {cur}": t["epi_100pct"],
            "LIMIT OUR SHARE USD": t["limit_our_share_usd"],
            "EPI OUR SHARE USD": t["epi_our_share_usd"],
            "Brokerage": t["brokerage_avg"], "Reins Tax": "", "Notes": "",
        })
    return pd.DataFrame(rows)


def build_uw_sheet_from_pack(
    pack,                              # ParsedRenewalPack
    shares_resolved: dict,             # output de determine_shares()
    treaty_metrics: list[dict],        # output de compute_treaty_metrics()
    header: TreatyHeader,
) -> UWSheet:
    """Construit un UWSheet à partir d'un renewal pack analysé + parts."""
    sheet = UWSheet(header=header)
    fx = header.exchange_rate_to_usd

    # Map treaty metrics par clé
    metrics_by_key = {m['lob_treaty'].lower().replace(' ', '_'): m
                       for m in treaty_metrics}

    for key, info in shares_resolved.items():
        if info["share_pct"] <= 0:
            continue
        t = info["treaty"]
        m = metrics_by_key.get(key.replace('_', ' ').replace(' ', '_'))

        epi_100 = info["prime_100_local"] / max(info["share_pct"], 1e-9) * info["share_pct"]
        # Recompute clean
        epi_100 = info["prime_100_local"]

        if m:
            comm = m.get('commission_pct', 0.30)
            lr_hist = m.get('lr_observed', 0)
            lr_est = m.get('lr_credible_buhlmann', 0)
        else:
            comm, lr_hist, lr_est = 0.30, 0, 0

        if t.treaty_type in ('qs', 'surplus', 'facility'):
            # Proportional
            ceded_limit = epi_100 * 5  # heuristic : 5x premium for capacity
            sheet.prop_lines.append(PropTreatyLine(
                treaty_name=f"{t.lob.title()} {t.treaty_type.upper()}",
                share_pct=info["share_pct"],
                leader=header.leader,
                limit_org_cur_100pct=ceded_limit * 1.5,
                ceded_limit_org_cur=ceded_limit,
                event_limit_org_cur=0,
                epi_org_cur=epi_100,
                historical_lr=lr_hist,
                estimated_lr=lr_est,
                commission_pct=comm,
                brokerage_pct=0.025,
            ))
        elif t.treaty_type == 'xol':
            sheet.nonprop_lines.append(NonPropTreatyLine(
                treaty_name=t.lob,
                share_pct=info["share_pct"],
                leader=header.leader,
                limit_org_cur_100pct=epi_100 * 10,
                deductible_org_cur=epi_100 * 2,
                reinstatement_str="2@100%",
                adjustable_rate_pct=0.05,
                gnpi_org_cur=epi_100 * 20,
                epi_org_cur=epi_100,
                brokerage_pct=0.10,
            ))

    return sheet


def add_xol_layers_from_pack(
    sheet: UWSheet, pack, share_pct: float = 0.10,
    layer_filter: str = "Layer 1",
) -> None:
    """Ajoute les couches XOL d'un pack à la section non-prop du sheet."""
    if pack.xol_layers.empty:
        return
    df = pack.xol_layers.copy()
    df = df.sort_values('uy').drop_duplicates(subset=['layer'], keep='last')
    for _, row in df.iterrows():
        layer_label = row['layer']
        if layer_filter and layer_filter not in layer_label:
            continue
        sheet.nonprop_lines.append(NonPropTreatyLine(
            treaty_name=f"Whole Account XOL {layer_label}",
            layer_n=int(layer_label.replace('Layer ', '')) if 'Layer ' in layer_label else 1,
            share_pct=share_pct,
            leader=sheet.header.leader,
            limit_org_cur_100pct=float(row['max_liability']),
            deductible_org_cur=float(row['priority']),
            reinstatement_str="2@100% (typical)",
            adjustable_rate_pct=(float(row['premium_total']) / float(row['gnpi_estimated'])
                                  if row['gnpi_estimated'] else 0),
            gnpi_org_cur=float(row['gnpi_estimated']),
            epi_org_cur=float(row['premium_total']),
            brokerage_pct=0.10,
            notes=f"UY {row['uy']} historical baseline",
        ))


def write_uw_sheet_to_excel(sheet: UWSheet, output_path: str,
                             include_clauses: bool = True) -> None:
    """Génère un .xlsx au format slip souscripteur professionnel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active
    ws.title = "UW Sheet"

    h = sheet.header
    cur = h.original_currency

    # Header
    ws['A1'] = f"REINSURANCE UNDERWRITING SHEET"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells('A1:N1')

    info_rows = [
        ("Company", h.company),
        ("Broker", h.broker),
        ("Inception Date", h.inception_date),
        ("Expiry Date", h.expiry_date),
        ("Period basis", h.period_basis),
        ("Leader", h.leader),
        ("Original Currency of Treaty", h.original_currency),
        ("Exchange rate to USD", f"{h.exchange_rate_to_usd:.4f}"),
        ("Business class", h.business_class),
        ("Country / Territory", h.country),
        ("Underwriter", h.underwriter),
        ("Quote ID", h.quote_id),
    ]
    for i, (k, v) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    row_offset = 3 + len(info_rows) + 2

    # PROP section
    if sheet.prop_lines:
        ws.cell(row=row_offset, column=1,
                value="PROPORTIONAL TREATIES (QS / Surplus / Facility)").font = \
            Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row_offset, column=1).fill = PatternFill("solid", fgColor="305496")
        ws.merge_cells(start_row=row_offset, start_column=1,
                        end_row=row_offset, end_column=16)
        row_offset += 1
        df_prop = prop_to_dataframe(sheet)
        for r_idx, row in enumerate(dataframe_to_rows(df_prop, index=False, header=True)):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_offset + r_idx, column=c_idx, value=val)
                if r_idx == 0:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="4472C4")
                elif df_prop.iloc[r_idx - 1, 0] == "TOTAL PROP":
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="D9E1F2")
        row_offset += len(df_prop) + 3

    # NON-PROP section
    if sheet.nonprop_lines:
        ws.cell(row=row_offset, column=1,
                value="NON-PROPORTIONAL TREATIES (XL Working / Cat / Whole Account)").font = \
            Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row_offset, column=1).fill = PatternFill("solid", fgColor="305496")
        ws.merge_cells(start_row=row_offset, start_column=1,
                        end_row=row_offset, end_column=16)
        row_offset += 1
        df_np = nonprop_to_dataframe(sheet)
        for r_idx, row in enumerate(dataframe_to_rows(df_np, index=False, header=True)):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_offset + r_idx, column=c_idx, value=val)
                if r_idx == 0:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="4472C4")
                elif df_np.iloc[r_idx - 1, 0] == "TOTAL NON PROP":
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="D9E1F2")
        row_offset += len(df_np) + 3

    # Grand total
    gt = sheet.grand_total()
    ws.cell(row=row_offset, column=1, value="GRAND TOTAL — OUR SHARE").font = \
        Font(size=12, bold=True, color="FFFFFF")
    ws.cell(row=row_offset, column=1).fill = PatternFill("solid", fgColor="C65911")
    ws.merge_cells(start_row=row_offset, start_column=1,
                    end_row=row_offset, end_column=4)
    row_offset += 1
    for k, label in [("total_epi_our_share_usd", "Total EPI OUR SHARE (USD)"),
                      ("total_limit_our_share_usd", "Total LIMIT OUR SHARE (USD)"),
                      ("expected_total_loss_usd", "Expected Total Loss (USD)")]:
        ws.cell(row=row_offset, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_offset, column=2, value=f"{gt[k]:,.0f}")
        row_offset += 1

    row_offset += 2

    # Clauses & wording
    if include_clauses:
        ws.cell(row=row_offset, column=1, value="WORDING / CLAUSES / CONDITIONS").font = \
            Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row_offset, column=1).fill = PatternFill("solid", fgColor="305496")
        ws.merge_cells(start_row=row_offset, start_column=1,
                        end_row=row_offset, end_column=16)
        row_offset += 1

        for section, items in [("CLAUSES", sheet.clauses),
                                ("EXCLUSIONS", sheet.exclusions),
                                ("WARRANTIES", sheet.warranties)]:
            if not items:
                continue
            ws.cell(row=row_offset, column=1, value=section).font = Font(bold=True)
            row_offset += 1
            for it in items:
                ws.cell(row=row_offset, column=1, value=f"• {it}")
                ws.merge_cells(start_row=row_offset, start_column=1,
                                end_row=row_offset, end_column=16)
                row_offset += 1
            row_offset += 1

    # Signatures
    row_offset += 2
    ws.cell(row=row_offset, column=1, value="SIGNATURES").font = \
        Font(size=11, bold=True, color="FFFFFF")
    ws.cell(row=row_offset, column=1).fill = PatternFill("solid", fgColor="305496")
    ws.merge_cells(start_row=row_offset, start_column=1,
                    end_row=row_offset, end_column=16)
    row_offset += 2
    sig_rows = [("Cedant", h.company),
                 ("Broker", h.broker),
                 ("Reinsurer", h.leader),
                 ("Underwriter", h.underwriter),
                 ("Date", datetime.utcnow().strftime("%Y-%m-%d"))]
    for label, val in sig_rows:
        ws.cell(row=row_offset, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_offset, column=2, value=val)
        ws.cell(row=row_offset, column=4, value="Signature: ____________________")
        row_offset += 1

    # Column widths
    widths = {1: 32, 2: 18, 3: 12, 4: 22, 5: 22, 6: 18, 7: 18, 8: 18, 9: 22,
              10: 22, 11: 14, 12: 14, 13: 14, 14: 14, 15: 18, 16: 30}
    for col, w in widths.items():
        ws.column_dimensions[chr(64 + col)].width = w

    wb.save(output_path)


def uw_sheet_to_html(sheet: UWSheet) -> str:
    """Génère un HTML formaté pour affichage Streamlit."""
    h = sheet.header
    cur = h.original_currency
    df_p = prop_to_dataframe(sheet)
    df_np = nonprop_to_dataframe(sheet)
    gt = sheet.grand_total()

    html = f"""
    <style>
        .uw-header {{ background:#1F4E78; color:white; padding:12px; font-size:18px; font-weight:bold; }}
        .uw-section {{ background:#305496; color:white; padding:8px; font-weight:bold; margin-top:18px; }}
        .uw-info {{ display:grid; grid-template-columns:200px 1fr; gap:4px 12px; padding:12px; }}
        .uw-info b {{ color:#333; }}
        table.uw {{ border-collapse:collapse; width:100%; font-size:12px; }}
        table.uw th {{ background:#4472C4; color:white; padding:6px; text-align:left; }}
        table.uw td {{ padding:5px 6px; border-bottom:1px solid #DDD; }}
        table.uw tr:nth-child(even) {{ background:#F8F9FA; }}
        .total-row td {{ font-weight:bold; background:#D9E1F2 !important; }}
        .grand-total {{ background:#FFF2CC; padding:12px; margin-top:18px; border:2px solid #C65911; font-weight:bold; }}
    </style>
    <div class="uw-header">REINSURANCE UNDERWRITING SHEET — {h.company}</div>
    <div class="uw-info">
        <b>Company</b><span>{h.company}</span>
        <b>Broker</b><span>{h.broker}</span>
        <b>Leader</b><span>{h.leader}</span>
        <b>Inception Date</b><span>{h.inception_date}</span>
        <b>Period basis</b><span>{h.period_basis}</span>
        <b>Original Currency</b><span>{h.original_currency}</span>
        <b>FX → USD</b><span>{h.exchange_rate_to_usd:.4f}</span>
        <b>Business class</b><span>{h.business_class}</span>
    </div>
    """

    if not df_p.empty:
        html += '<div class="uw-section">PROPORTIONAL TREATIES</div>'
        html += df_p.to_html(classes='uw', index=False, escape=False,
                              float_format=lambda x: f"{x:,.2f}" if isinstance(x, float) else x)
    if not df_np.empty:
        html += '<div class="uw-section">NON-PROPORTIONAL TREATIES</div>'
        html += df_np.to_html(classes='uw', index=False, escape=False,
                               float_format=lambda x: f"{x:,.2f}" if isinstance(x, float) else x)

    html += f"""
    <div class="grand-total">
        <div style="font-size:14px;color:#C65911;">GRAND TOTAL — OUR SHARE</div>
        Total EPI OUR SHARE: <b>USD {gt['total_epi_our_share_usd']:,.0f}</b><br>
        Total LIMIT OUR SHARE: <b>USD {gt['total_limit_our_share_usd']:,.0f}</b><br>
        Expected Total Loss: <b>USD {gt['expected_total_loss_usd']:,.0f}</b>
    </div>
    """
    return html
