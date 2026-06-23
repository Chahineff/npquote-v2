"""P2.6 — Parser générique de renewal packs courtiers.

Détection automatique de la structure :
- Identification des LOB (Fire / Engineering / GA / Motor / Marine / etc.)
- Détection des feuilles EPI, Statistics, XOL, Triangulation, Major Losses
- Mapping flexible des colonnes (UY, Premium, Commission, Incurred, LR)
- Multi-format : .xlsx, .xlsm, .csv

Compatible avec les templates Aon, GuyCarp, Willis Re, Lockton Re,
Marsh, Howden Re, et les templates cédantes du Moyen-Orient / Afrique.
"""
from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
import re
import openpyxl
import pandas as pd
import warnings
warnings.filterwarnings('ignore')


# Mots-clés pour détection des feuilles
SHEET_PATTERNS = {
    "epi":         ["epi", "premium income", "gnpi", "gpi", "estimated premium",
                     "gross written"],
    "statistics":  ["statistics", "statistical", "stat", "loss exp", "experience"],
    "triangulation": ["triangulation", "triangle", "development"],
    "xol":         ["xol", "xl", "excess", "non.prop"],
    "risk_profile": ["risk profile", "profile", "portfolio profile"],
    "major_losses": ["major loss", "large loss", "claim list", "loss list",
                      "losses above"],
    "large_risks": ["large risk", "biggest", "top exposure"],
    "cresta":       ["cresta", "natcat"],
}

# Mots-clés pour LOB
LOB_PATTERNS = {
    "fire":          ["fire", "property", "fid", "fls"],
    "engineering":   ["engineering", "car", "ear", "epp", "machinery",
                       "mbd", "cpm", "ee"],
    "ga":            ["general accident", " ga ", "ga ", "personal accident"],
    "motor":         ["motor", "auto"],
    "marine":        ["marine", "hull", "cargo"],
    "aviation":      ["aviation"],
    "energy":        ["energy", "onshore", "offshore"],
    "casualty":      ["casualty", "liability", "professional indemnity"],
    "health_life":   ["health", "medical", "life"],
}


def _normalize(s: str) -> str:
    """Normalise un nom : lower, sans accents, ponctuation réduite."""
    if s is None:
        return ""
    s = str(s).lower().strip()
    s = re.sub(r'[éèêë]', 'e', s)
    s = re.sub(r'[àâä]', 'a', s)
    s = re.sub(r'[^a-z0-9 _-]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def detect_lob_from_text(text: str) -> str | None:
    """Détecte le LOB à partir d'un texte (nom de fichier ou de feuille)."""
    txt = _normalize(text)
    for lob, patterns in LOB_PATTERNS.items():
        for p in patterns:
            if p in txt:
                return lob
    return None


def detect_sheet_type(sheet_name: str) -> str | None:
    """Détecte le type de feuille (EPI/stats/XOL/etc.)."""
    s = _normalize(sheet_name)
    # Order matters: xol must be matched before statistics (XL Statistics)
    for stype, patterns in [(t, SHEET_PATTERNS[t])
                             for t in ['xol', 'triangulation', 'major_losses',
                                        'large_risks', 'cresta', 'risk_profile',
                                        'statistics', 'epi']]:
        for p in patterns:
            if p in s:
                return stype
    return None


@dataclass
class DetectedSheet:
    name: str
    sheet_type: str | None
    lob: str | None
    n_rows: int
    n_cols: int


@dataclass
class DetectedTreaty:
    """Un traité détecté dans le renewal pack."""
    lob: str                # fire / engineering / ga / motor / ...
    treaty_type: str         # qs / surplus / facility / xol
    by_uy: pd.DataFrame      # série historique
    metadata: dict = field(default_factory=dict)


def scan_folder(folder: Path, currency_hint: str = None) -> dict:
    """Scanne un dossier de renewal pack et retourne sa structure détectée.

    Pour chaque fichier .xlsx/.xlsm trouvé :
      - Détecte les feuilles et leur type
      - Détecte le LOB depuis le nom de fichier OU les noms de feuilles
      - Extrait l'EPI / statistiques / XOL
    """
    folder = Path(folder)
    if not folder.exists():
        raise FileNotFoundError(folder)

    files = list(folder.glob("*.xlsx")) + list(folder.glob("*.xlsm"))
    detected = {
        "folder": str(folder),
        "n_files": len(files),
        "files": {},
        "treaties": [],
        "currency_detected": currency_hint,
        "warnings": [],
    }

    for f in files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        except Exception as e:
            detected["warnings"].append(f"Cannot read {f.name}: {e}")
            continue

        file_lob_hint = detect_lob_from_text(f.stem)
        file_info = {"path": str(f), "lob_hint": file_lob_hint, "sheets": []}

        for sn in wb.sheetnames:
            ws = wb[sn]
            try:
                stype = detect_sheet_type(sn)
                sheet_lob = detect_lob_from_text(sn) or file_lob_hint
                rows = getattr(ws, 'max_row', 0) or 0
                cols = getattr(ws, 'max_column', 0) or 0
                file_info["sheets"].append({
                    "name": sn, "type": stype, "lob": sheet_lob,
                    "rows": rows, "cols": cols,
                })
            except Exception:
                continue
        detected["files"][f.name] = file_info

    return detected


def extract_stats_block(ws, treaty_type_kw: str, lob: str
                        ) -> pd.DataFrame:
    """Extrait un bloc Statistics : trouve l'en-tête puis itère les lignes UY.

    treaty_type_kw : 'qs' / 'surplus' / 'facility'
    """
    rows = []
    in_block = False
    header_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str):
            vn = _normalize(v)
            if vn == treaty_type_kw or vn.startswith(treaty_type_kw + ' '):
                in_block = True
                continue
            if in_block and vn in ('total', 'qs', 'surplus', 'facility') \
               and vn != treaty_type_kw:
                in_block = False
                continue
            if in_block and ('uw/y' in vn or 'uw y' in vn or vn.startswith('uw')):
                header_row = r
                continue
        if in_block and header_row and r > header_row + 1:
            cells = [ws.cell(row=r, column=c).value for c in range(1, 9)]
            uy_v = cells[0]
            if isinstance(uy_v, str) and any(ch.isdigit() for ch in uy_v):
                try:
                    rows.append({
                        "uy": str(uy_v).strip(),
                        "premium": _to_float(cells[1]),
                        "commission": _to_float(cells[2]),
                        "tax": _to_float(cells[3]),
                        "paid": _to_float(cells[4]),
                        "outstanding": _to_float(cells[5]),
                        "incurred": _to_float(cells[6]),
                        "loss_ratio": _to_float(cells[7]),
                    })
                except (TypeError, ValueError):
                    continue
    return pd.DataFrame(rows)


def _to_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace(" ", ""))
    except (TypeError, ValueError):
        return 0.0


def extract_epi_block(ws) -> dict:
    """Extrait l'EPI multi-état (actual / estimated / revised / next).

    Cherche les sections "Actual Premium", "Estimated Premium",
    "Revised EPI" et "EPI UW/Y next".
    """
    states_found = {}
    keywords_map = {
        "actual":    ["actual premium"],
        "estimated": ["estimated premium"],
        "revised":   ["revised epi", "revised premium"],
        "next":      ["epi uw", "epi  uw"],
    }
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if not isinstance(v, str):
            continue
        vn = _normalize(v)
        for state, kws in keywords_map.items():
            if any(kw in vn for kw in kws) and state not in states_found:
                states_found[state] = r
                break

    result = {}
    for state, start_r in states_found.items():
        total_row = None
        for offset in range(0, 6):
            v2 = ws.cell(row=start_r + offset, column=2).value
            if isinstance(v2, str) and 'total' in v2.lower():
                total_row = start_r + offset
                break
        if total_row is None:
            for offset in range(0, 6):
                if any(isinstance(ws.cell(row=start_r + offset, column=c).value,
                                   (int, float))
                       for c in range(3, 9)):
                    total_row = start_r + offset
                    break
        if total_row:
            result[state] = {
                "total":          _to_float(ws.cell(row=total_row, column=3).value),
                "qs_retention":   _to_float(ws.cell(row=total_row, column=4).value),
                "qs_cession":     _to_float(ws.cell(row=total_row, column=5).value),
                "surplus":        _to_float(ws.cell(row=total_row, column=6).value),
                "facility":       _to_float(ws.cell(row=total_row, column=7).value),
                "fac_outwards":   _to_float(ws.cell(row=total_row, column=8).value),
            }
    return result


def extract_xol_block(ws) -> pd.DataFrame:
    """Extrait les couches XOL depuis une feuille Statistics."""
    rows = []
    current_uy = None
    for r in range(1, ws.max_row + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, 14)]
        for v in cells:
            if isinstance(v, str) and v.startswith('UY '):
                current_uy = v.replace('UY ', '').strip()
                break
        v0 = cells[0]
        if isinstance(v0, str) and v0.lower().startswith('layer ') and current_uy:
            if cells[1] is not None:
                rows.append({
                    "uy": current_uy, "layer": v0,
                    "max_liability": _to_float(cells[1]),
                    "priority": _to_float(cells[2]),
                    "gnpi_estimated": _to_float(cells[3]),
                    "gnpi_realized": _to_float(cells[4]),
                    "premium_md": _to_float(cells[5]),
                    "premium_adj": _to_float(cells[6]),
                    "premium_total": _to_float(cells[7]),
                    "paid_losses": _to_float(cells[8]),
                    "os_losses": _to_float(cells[9]),
                    "incurred": _to_float(cells[10]),
                })
    return pd.DataFrame(rows)


@dataclass
class ParsedRenewalPack:
    """Pack complet parsé."""
    folder: str
    cedante: str
    currency: str
    treaties: list                # DetectedTreaty
    epi_by_lob: dict
    xol_layers: pd.DataFrame
    detection_report: dict
    warnings: list = field(default_factory=list)


def parse_renewal_folder(
    folder: Path,
    cedante: str = "Unknown",
    currency: str = "EUR",
) -> ParsedRenewalPack:
    """Parse complet d'un dossier renewal pack.

    Pipeline :
      1. Scan : identifie fichiers, feuilles, LOB
      2. Extract : pour chaque fichier, parse EPI + Stats + XOL si trouvés
      3. Assemble : structure unifiée ParsedRenewalPack
    """
    folder = Path(folder)
    detection = scan_folder(folder, currency_hint=currency)

    treaties = []
    epi_by_lob = {}
    xol_all = pd.DataFrame()
    warnings_acc = []

    for fname, finfo in detection["files"].items():
        path = Path(finfo["path"])
        file_lob = finfo["lob_hint"]
        wb = openpyxl.load_workbook(path, data_only=True)

        for sheet_info in finfo["sheets"]:
            sn = sheet_info["name"]
            stype = sheet_info["type"]
            lob = sheet_info["lob"] or file_lob or "unknown"
            if stype is None or sheet_info["rows"] == 0:
                continue
            ws = wb[sn]

            if stype == "epi":
                epi = extract_epi_block(ws)
                if epi and lob != "unknown":
                    epi_by_lob[lob] = epi
            elif stype == "statistics":
                for tkw in ("qs", "surplus", "facility"):
                    df = extract_stats_block(ws, tkw, lob)
                    if not df.empty:
                        treaties.append(DetectedTreaty(
                            lob=lob, treaty_type=tkw, by_uy=df,
                            metadata={"file": fname, "sheet": sn},
                        ))
            elif stype == "xol":
                df = extract_xol_block(ws)
                if not df.empty:
                    df['file'] = fname
                    xol_all = pd.concat([xol_all, df], ignore_index=True)

    if not treaties:
        warnings_acc.append("Aucun traité QS/Surplus/Facility détecté")
    if epi_by_lob == {}:
        warnings_acc.append("Aucun EPI extrait — vérifier structure des fichiers")

    return ParsedRenewalPack(
        folder=str(folder), cedante=cedante, currency=currency,
        treaties=treaties, epi_by_lob=epi_by_lob,
        xol_layers=xol_all, detection_report=detection,
        warnings=warnings_acc,
    )
